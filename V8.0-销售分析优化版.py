# -*- coding: utf-8 -*-
"""
供应商报表生成器 - v8.0
1.  版本合并：基于 v7.0 (移除缓存) 版本，并整合 v7.1 的核心功能。
2.  报表求和优化：采用 v7.1 的 SUBTOTAL 函数直接写入单元格的方式，以增强在不同Excel版本中的兼容性和稳定性。
3.  兼容性增强保存：集成 v7.1 的保存机制，在 openpyxl 保存后，尝试调用 WPS(ket.Application) 或 Excel(Excel.Application) 的COM组件重新保存文件，旨在修复文件在移动端（如微信、钉钉）可能出现的格式或图表显示问题。
4.  用户体验优化：保留 v7.0 的处理进度显示、界面响应性优化和错误提示。
5.  功能扩展：保留 v7.0 的数据质量检查、CSV导出支持和可视化图表功能。
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import os
import calendar
import threading
import re
import subprocess
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.chart import BarChart, Reference, LineChart

# ==================== 配置管理 ====================
class Config:
    """配置管理类"""
    FOLDERS = {'data': '主数据', 'reports': '报表'}
    STD_COLS = {
        'BRAND': '品牌', 'BARCODE': '条码', 'NAME': '名称', 'SPEC': '规格', 'PRICE': '定价',
        'STOCK': '库存量', 'LAST_INBOUND_DATE': '最后进货日', 'REMARK': '备注',
        'TOTAL_REVENUE': '总实收', 'TOTAL_ORDERS': '总笔数', 'TOTAL_SALES_QTY': '总销量',
        'WEEK_PERIOD': '周度期间', 'SALES_TIME': '销售时间', 'SALES_QTY': '销售数量',
        'REVENUE': '实收金额', 'ORDER_ID': '流水号'
    }
    COLUMN_MAPPINGS = {
        'brand': ['商品品牌', '品牌', 'Brand'],
        'barcode': ['商品条码', '条码', 'Barcode'],
        'name': ['商品名称', '名称（必填）', '名称', '商品名', 'Name'],
        'spec': ['规格', '商品规格', '规格型号'],
        'price': ['销售价（必填）', '销售价', '定价', '商品原价'],
        'received_qty': ['实收量'],
        'flow_qty': ['货流量'],
        'check_diff': ['差异库存']
    }
    FILE_PATTERNS = {
        'product': '商品资料',
        'sales': 'sales_data.xlsx',
        'inventory_flow': 'inventory_flow_data.xlsx',
        'inventory_check': '盘点盈亏明细.xlsx'
    }
    DATE_COLUMNS = {
        'flow': ['下单时间', '日期', '进货时间', '入库时间'],
        'check': ['盘点时间', '日期'],
        'sales': ['销售时间']
    }

    @staticmethod
    def get_file_path(file_type):
        pattern = Config.FILE_PATTERNS.get(file_type, '')
        if file_type == 'product':
            return DataProcessor.find_file_in_data_folder(pattern)
        return os.path.join(Config.FOLDERS['data'], pattern)

    @staticmethod
    def get_report_path(filename):
        return os.path.join(Config.FOLDERS['reports'], filename)

    @staticmethod
    def ensure_folders():
        for folder in Config.FOLDERS.values():
            if not os.path.exists(folder):
                os.makedirs(folder)
                print(f"📁 创建文件夹: {folder}")

# ==================== 缓存管理已移除 ====================
# 根据用户要求，移除缓存机制，直接从主数据文件夹读取数据

# ==================== 通用工具 ====================
EXCEL_FORMATS = {
    'currency': '_* #,##0.00_ ;-* #,##0.00_ ;_* \"-\"??_ ;_-@_',
    'integer': '#,##0_ ;-#,##0',
    'date': 'yyyy-mm-dd'
}

def open_file_or_folder(path):
    """跨平台打开文件或文件夹"""
    try:
        if not os.path.exists(path):
            messagebox.showerror("错误", f"路径不存在: {path}")
            return
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("错误", f"无法打开路径: {e}")

# ==================== 数据处理 ====================
class DataProcessor:
    """数据加载、清洗和列名查找

    该类负责从Excel文件加载数据，进行清洗和列映射。
    """
    def __init__(self) -> None:
        pass  # 移除缓存管理器
    
    @staticmethod
    def clean_numeric_column(series: pd.Series, remove_chars: list[str] = None) -> pd.Series:
        """清洗数值列：移除指定字符并转换为数值类型

        Args:
            series (pd.Series): 输入序列
            remove_chars (list[str], optional): 要移除的字符列表. Defaults to None.

        Returns:
            pd.Series: 清洗后的数值序列
        """
        remove_chars = remove_chars or ['￥', '¥', '$', ',']
        if series is None or series.empty:
            return pd.Series(dtype=float)
        cleaned = pd.to_numeric(series, errors='coerce')
        if cleaned.isna().sum() > 0:
            s_str = series.astype(str)
            for char in remove_chars:
                s_str = s_str.str.replace(char, '', regex=False)
            cleaned = pd.to_numeric(s_str, errors='coerce')
        return cleaned.fillna(0)

    @staticmethod
    def find_column(df, possible_names):
        if df is None or df.empty:
            return None
        for name in possible_names:
            if name in df.columns:
                return name
        return None

    def load_excel_with_mapping(self, file_path_or_pattern, dtype_mapping=None, chunked=False):
        # 直接从文件加载，不使用缓存
        file_path = file_path_or_pattern
        if not os.path.isabs(file_path) and not os.path.exists(file_path):
            full_path = os.path.join(Config.FOLDERS['data'], file_path)
            if os.path.exists(full_path):
                file_path = full_path
        if not os.path.exists(file_path):
            print(f"⚠️ 文件不存在: {file_path}")
            return pd.DataFrame()
        try:
            if chunked and os.path.getsize(file_path) > 50 * 1024 * 1024:  # >50MB
                df = self.load_excel_chunked(file_path, dtype_mapping)
            else:
                df = pd.read_excel(file_path, dtype=dtype_mapping or {})
            print(f"✅ 成功加载: {os.path.basename(file_path)} ({len(df)} 条记录)")
            return df
        except Exception as e:
            messagebox.showerror("文件加载错误", f"加载文件 '{os.path.basename(file_path)}' 时出错:\n{e}")
            return pd.DataFrame()

    @staticmethod
    def load_excel_chunked(file_path, dtype_mapping=None, chunk_size=50000):
        """分块读取大文件"""
        chunks = []
        print(f"📦 文件大于50MB，开始分块加载: {os.path.basename(file_path)}")
        for i, chunk in enumerate(pd.read_excel(file_path, chunksize=chunk_size, dtype=dtype_mapping or {})):
            print(f"  - 加载块 {i+1}...")
            chunks.append(chunk)
        df = pd.concat(chunks, ignore_index=True)
        print(f"✅ 分块加载完成: {len(df)} 条记录")
        return df

    @staticmethod
    def find_file_in_data_folder(pattern):
        data_folder = Config.FOLDERS['data']
        if not os.path.exists(data_folder):
            return None
        try:
            files = [(os.path.join(data_folder, f), os.path.getmtime(os.path.join(data_folder, f)))
                     for f in os.listdir(data_folder)
                     if pattern in f and f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
            if files:
                files.sort(key=lambda x: x[1], reverse=True)
                return files[0][0]
            return None
        except Exception as e:
            print(f"❌ 查找文件失败 ({pattern}): {e}")
            return None

class InventoryCalculator:
    def __init__(self, data_processor):
        self.data_processor = data_processor

    # V7.0 MODIFIED: 移除缓存机制，直接计算
    def calculate_inventory(self, product_barcodes, flow_df, check_df, sales_df, end_date=None):
        print(f"📦 正在计算库存... (截止日期: {end_date.strftime('%Y-%m-%d') if end_date else '无'})")
        C = Config.STD_COLS

        # V6.5 MODIFIED: 如果提供了截止日期，则筛选所有相关数据
        if end_date:
            end_date_inclusive = datetime.combine(end_date, datetime.max.time())
            if not sales_df.empty and C['SALES_TIME'] in sales_df.columns:
                sales_df = sales_df[sales_df[C['SALES_TIME']] <= end_date_inclusive].copy()
            if not flow_df.empty and '日期' in flow_df.columns:
                flow_df = flow_df[flow_df['日期'] <= end_date_inclusive].copy()
            if not check_df.empty and '日期' in check_df.columns:
                check_df = check_df[check_df['日期'] <= end_date_inclusive].copy()

        flow_s = flow_df.groupby(C['BARCODE'])['库存变动量'].sum() if not flow_df.empty else pd.Series(name='库存变动量')
        sales_s = sales_df.groupby(C['BARCODE'])[C['SALES_QTY']].sum() if not sales_df.empty else pd.Series(name=C['SALES_QTY'])
        check_s = check_df.groupby(C['BARCODE'])['差异库存'].sum() if not check_df.empty else pd.Series(name='差异库存')
        
        inventory_df = pd.DataFrame({C['BARCODE']: product_barcodes}).set_index(C['BARCODE'])
        inventory_df = inventory_df.join(flow_s).join(sales_s).join(check_s).fillna(0).reset_index()

        inventory_df[C['STOCK']] = (inventory_df['库存变动量'] - inventory_df[C['SALES_QTY']] + inventory_df['差异库存']).astype(int)
        
        last_inbound_info = self._get_last_inbound_info(flow_df, product_barcodes)
        inventory_df = inventory_df.merge(last_inbound_info, on=C['BARCODE'], how='left')
        inventory_df[C['REMARK']] = inventory_df.apply(lambda row: self._get_remark(row, flow_df), axis=1).fillna('')
        
        print(f"✅ 库存计算完成: {len(inventory_df)} 个商品")
        result = inventory_df[[C['BARCODE'], C['STOCK'], C['LAST_INBOUND_DATE'], C['REMARK']]]
        return result

    def _get_last_inbound_info(self, flow_df, product_barcodes):
        C = Config.STD_COLS
        if flow_df.empty:
            return pd.DataFrame({C['BARCODE']: product_barcodes, C['LAST_INBOUND_DATE']: ''})
        inbound_records = flow_df[flow_df['库存变动量'] > 0].copy()
        if inbound_records.empty:
            return pd.DataFrame({C['BARCODE']: product_barcodes, C['LAST_INBOUND_DATE']: ''})
        inbound_records['日期'] = pd.to_datetime(inbound_records['日期'])
        last_inbound = inbound_records.loc[inbound_records.groupby(C['BARCODE'])['日期'].idxmax()]
        last_inbound[C['LAST_INBOUND_DATE']] = last_inbound.apply(
            lambda r: f"{r['日期'].strftime('%Y-%m-%d')} ({int(r['库存变动量'])}件)", axis=1)
        return last_inbound[[C['BARCODE'], C['LAST_INBOUND_DATE']]]

    def _get_remark(self, row, flow_df) -> str:
        """获取备注：检查商品是否可能已退库

        Args:
            row (pd.Series): 当前行数据
            flow_df (pd.DataFrame): 库存流动数据

        Returns:
            str: 备注信息
        """
        C = Config.STD_COLS
        if flow_df.empty or row[C['STOCK']] > 0:
            return ''
        barcode_flow = flow_df[flow_df[C['BARCODE']] == row[C['BARCODE']]]
        if barcode_flow.empty:
            return ''
        last_record = barcode_flow.sort_values('日期').iloc[-1]
        if last_record.get('库存变动量', 0) < 0:
            return '商品可能已退库'
        return ''

class SalesAnalyzer:
    def __init__(self, data_processor):
        self.data_processor = data_processor

    # V7.0 MODIFIED: 移除缓存机制，直接分析
    def analyze_sales(self, filtered_sales, product_barcodes, week_periods):
        print("💰 正在按周分析销售数据...")
        C = Config.STD_COLS
        if filtered_sales.empty:
            return self._create_empty_sales_result(product_barcodes, week_periods)

        def get_period_label(sale_date):
            for start, end in week_periods:
                if start.date() <= sale_date.date() <= end.date():
                    return f"{start.month}.{start.day}-{end.month}.{end.day}"
            return None

        filtered_sales[C['WEEK_PERIOD']] = filtered_sales[C['SALES_TIME']].apply(get_period_label)

        sales_summary = filtered_sales.groupby(C['BARCODE']).agg({
            C['REVENUE']: 'sum',
            C['ORDER_ID']: 'nunique',
            C['SALES_QTY']: 'sum'
        }).rename(columns={
            C['REVENUE']: C['TOTAL_REVENUE'],
            C['ORDER_ID']: C['TOTAL_ORDERS'],
            C['SALES_QTY']: C['TOTAL_SALES_QTY']
        })

        weekly_sales_pivot = filtered_sales.pivot_table(
            index=C['BARCODE'],
            columns=C['WEEK_PERIOD'],
            values=C['SALES_QTY'],
            aggfunc='sum',
            fill_value=0
        )
        
        week_labels = [f"{start.month}.{start.day}-{end.month}.{end.day}" for start, end in week_periods]
        weekly_sales_pivot = weekly_sales_pivot.reindex(columns=week_labels, fill_value=0)

        all_products_sales = pd.DataFrame({C['BARCODE']: product_barcodes}).set_index(C['BARCODE']).join(
            sales_summary).join(weekly_sales_pivot).fillna(0).reset_index()
        print(f"✅ 周度销售分析完成: {len(all_products_sales)} 个商品")
        return all_products_sales

    def _create_empty_sales_result(self, product_barcodes, week_periods):
        C = Config.STD_COLS
        result = pd.DataFrame({
            C['BARCODE']: product_barcodes,
            C['TOTAL_REVENUE']: 0,
            C['TOTAL_ORDERS']: 0,
            C['TOTAL_SALES_QTY']: 0
        })
        week_labels = [f"{start.month}.{start.day}-{end.month}.{end.day}" for start, end in week_periods]
        for label in week_labels:
            result[label] = 0
        return result

class ProductManager:
    def __init__(self, data_processor):
        self.data_processor = data_processor

    def load_and_prep_data(self):
        # 直接从文件加载数据，不使用缓存
        product_file_path = Config.get_file_path('product')
        product_df = self.data_processor.load_excel_with_mapping(product_file_path)
        sales_df = self.data_processor.load_excel_with_mapping(
            Config.FILE_PATTERNS['sales'], chunked=True)
        flow_df = self.data_processor.load_excel_with_mapping(
            Config.FILE_PATTERNS['inventory_flow'], dtype_mapping={'商品条码': str, '条码': str})
        check_df = self.data_processor.load_excel_with_mapping(
            Config.FILE_PATTERNS['inventory_check'], dtype_mapping={'商品条码': str})
        sales_df = self._prep_sales_df(sales_df)
        flow_df = self._prep_flow_df(flow_df)
        check_df = self._prep_check_df(check_df)
        return {
            'product': product_df,
            'sales': sales_df,
            'inventory_flow': flow_df,
            'inventory_check': check_df
        }, product_file_path

    def _prep_sales_df(self, df):
        if df.empty:
            return df
        C, M, D = Config.STD_COLS, Config.COLUMN_MAPPINGS, Config.DATE_COLUMNS
        rename_map = {
            self.data_processor.find_column(df, M['barcode']): C['BARCODE'],
            self.data_processor.find_column(df, M['brand']): C['BRAND'],
            self.data_processor.find_column(df, D['sales']): C['SALES_TIME'],
            '实收金额': C['REVENUE'],
            '销售数量': C['SALES_QTY'],
            '流水号': C['ORDER_ID']
        }
        df.rename(columns={k: v for k, v in rename_map.items() if k}, inplace=True)
        df[C['SALES_TIME']] = pd.to_datetime(df.get(C['SALES_TIME']), errors='coerce')
        df[C['REVENUE']] = self.data_processor.clean_numeric_column(df.get(C['REVENUE']))
        df[C['SALES_QTY']] = self.data_processor.clean_numeric_column(df.get(C['SALES_QTY']))
        df.dropna(subset=[C['SALES_TIME'], C['BARCODE']], inplace=True)
        df[C['BARCODE']] = df[C['BARCODE']].astype(str)
        return df

    def _prep_flow_df(self, df):
        if df.empty:
            return df
        C, M, D = Config.STD_COLS, Config.COLUMN_MAPPINGS, Config.DATE_COLUMNS
        df.rename(columns={self.data_processor.find_column(df, M['barcode']): C['BARCODE']}, inplace=True)
        date_col = self.data_processor.find_column(df, D['flow'])
        if date_col:
            df.rename(columns={date_col: '日期'}, inplace=True)
            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
        if '实收量' in df.columns and '货流量' in df.columns:
            # 处理退货单的实收量格式："-", "0", "0.00" 等都视为无效值
            received_qty_raw = df['实收量'].astype(str).str.strip()
            # 将 "-", "0", "0.00", 空字符串等视为无效值
            invalid_values = ['-', '0', '0.00', '', 'nan', 'NaN', 'None']
            received_qty_clean = received_qty_raw.replace(invalid_values, np.nan)
            received_qty = pd.to_numeric(received_qty_clean, errors='coerce')
            flow_qty = pd.to_numeric(df['货流量'], errors='coerce').fillna(0)
            # 当实收量为无效值时，使用负的货流量（退货减库存）
            df['库存变动量'] = np.where(received_qty.isna() | (received_qty == 0), -flow_qty, received_qty)
        else:
            df['库存变动量'] = 0
        df.dropna(subset=[C['BARCODE']], inplace=True)
        df[C['BARCODE']] = df[C['BARCODE']].astype(str)
        return df

    def _prep_check_df(self, df):
        if df.empty:
            return df
        C, M, D = Config.STD_COLS, Config.COLUMN_MAPPINGS, Config.DATE_COLUMNS
        df.rename(columns={self.data_processor.find_column(df, M['barcode']): C['BARCODE']}, inplace=True)
        date_col = self.data_processor.find_column(df, D['check'])
        if date_col:
            df.rename(columns={date_col: '日期'}, inplace=True)
            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
        df['差异库存'] = pd.to_numeric(df.get('差异库存'), errors='coerce').fillna(0)
        df.dropna(subset=[C['BARCODE']], inplace=True)
        df[C['BARCODE']] = df[C['BARCODE']].astype(str)
        return df

    def get_all_brands(self, product_df, sales_df):
        # 关键修正：品牌数据仅从商品资料文件获取，避免销售数据中的旧品牌名造成干扰
        brands = set()
        C, M = Config.STD_COLS, Config.COLUMN_MAPPINGS
        if not product_df.empty:
            brand_col = self.data_processor.find_column(product_df, M['brand'])
            if brand_col:
                brands.update(product_df[brand_col].dropna().unique())
        # 过滤掉空值或仅包含空格的品牌
        return sorted([b for b in brands if pd.notna(b) and str(b).strip()])

    def build_master_product_data(self, product_df, selected_brands):
        C, M = Config.STD_COLS, Config.COLUMN_MAPPINGS
        if product_df.empty:
            messagebox.showwarning("警告", "商品资料文件为空，报表将缺少名称、规格和定价信息。")
            return pd.DataFrame(columns=[C['BRAND'], C['BARCODE'], C['NAME'], C['SPEC'], C['PRICE']])
        b_col = self.data_processor.find_column(product_df, M['brand'])
        bc_col = self.data_processor.find_column(product_df, M['barcode'])
        n_col = self.data_processor.find_column(product_df, M['name'])
        if not all([b_col, bc_col, n_col]):
            messagebox.showerror("错误", "商品资料文件缺少必要列（品牌、条码、名称）。")
            return pd.DataFrame()
        master_data = product_df[product_df[b_col].isin(selected_brands)].copy()
        result = pd.DataFrame()
        result[C['BRAND']] = master_data[b_col]
        result[C['BARCODE']] = master_data[bc_col].astype(str)
        result[C['NAME']] = master_data[n_col]
        spec_col = self.data_processor.find_column(product_df, M['spec'])
        result[C['SPEC']] = master_data[spec_col] if spec_col is not None else ''
        price_col = self.data_processor.find_column(product_df, M['price'])
        result[C['PRICE']] = self.data_processor.clean_numeric_column(master_data.get(price_col)) if price_col is not None else 0
        result.drop_duplicates(subset=[C['BARCODE']], inplace=True)
        print(f"✅ 商品主数据构建完成: {len(result)} 个商品")
        return result.reset_index(drop=True)

# ==================== 数据质量检查 ====================
class DataQualityChecker:
    """数据质量检查工具"""
    
    @staticmethod
    def check_data_quality(data_frames):
        """检查数据质量"""
        issues = []
        C = Config.STD_COLS
        
        # 检查销售数据
        sales_df = data_frames.get('sales', pd.DataFrame())
        if not sales_df.empty:
            # 移除了对负销售额和负销售数量的检查，因为退货是正常业务场景
            
            # 检查时间范围
            if C['SALES_TIME'] in sales_df.columns:
                min_date = sales_df[C['SALES_TIME']].min()
                max_date = sales_df[C['SALES_TIME']].max()
                if (max_date - min_date).days > 365 * 2:
                    issues.append(f"销售数据时间跨度较长 ({min_date.date()} 到 {max_date.date()})，可能影响性能")
        
        # 检查库存流动数据
        flow_df = data_frames.get('inventory_flow', pd.DataFrame())
        if not flow_df.empty:
            # 检查零库存变动记录
            zero_flow = flow_df[flow_df['库存变动量'] == 0]
            if not zero_flow.empty:
                issues.append(f"库存流动数据中发现 {len(zero_flow)} 条零库存变动记录")
        
        return issues

# ==================== 报表生成器 ====================
class ReportGenerator:
    def __init__(self, data_processor, inventory_calc, sales_analyzer, product_manager):
        self.data_processor = data_processor
        self.inventory_calc = inventory_calc
        self.sales_analyzer = sales_analyzer
        self.product_manager = product_manager
        self.progress_callback = None

    def set_progress_callback(self, callback):
        """设置进度回调函数"""
        self.progress_callback = callback

    def generate_report(self, data_frames, selected_brands, start_date, end_date, sort_params, export_format='excel'):
        print("🔄 开始生成报表...")
        if self.progress_callback:
            self.progress_callback(0, "开始生成报表...")
        C = Config.STD_COLS
        product_df = data_frames['product']
        sales_df = data_frames['sales'] # This is the full, unprepared sales_df
        flow_df = data_frames['inventory_flow']
        check_df = data_frames['inventory_check']

        if self.progress_callback:
            self.progress_callback(5, "构建商品主数据...")
        master_products = self.product_manager.build_master_product_data(product_df, selected_brands)
        if master_products.empty:
            messagebox.showerror("错误", "无任何有效的商品数据。")
            return False, None

        end_date_inclusive = datetime.combine(end_date, datetime.max.time())
        start_date_inclusive = datetime.combine(start_date, datetime.min.time())
        selected_barcodes = master_products[C['BARCODE']].astype(str).unique()

        if self.progress_callback:
            self.progress_callback(10, f"筛选销售时段: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        # Filter sales for the selected date range for reporting purposes
        filtered_sales = sales_df[
            (sales_df[C['BARCODE']].astype(str).isin(selected_barcodes)) &
            (sales_df[C['SALES_TIME']] >= start_date_inclusive) &
            (sales_df[C['SALES_TIME']] <= end_date_inclusive)
        ].copy()
        print(f"📋 筛选时段: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}, 共 {len(filtered_sales)} 条记录")

        week_periods = self._get_week_periods(start_date, end_date)
        week_labels = [f"{start.month}.{start.day}-{end.month}.{end.day}" for start, end in week_periods]

        # V6.5: Calculate final inventory for the main report (up to the latest data)
        if self.progress_callback:
            self.progress_callback(20, "计算库存数据...")
        inventory_data = self.inventory_calc.calculate_inventory(
            master_products[C['BARCODE']].tolist(), flow_df, check_df, sales_df
        )
        if self.progress_callback:
            self.progress_callback(40, "分析销售数据...")
        sales_data = self.sales_analyzer.analyze_sales(filtered_sales, master_products[C['BARCODE']].tolist(), week_periods)

        if self.progress_callback:
            self.progress_callback(70, "合并数据...")
        final_data = master_products.merge(inventory_data, on=C['BARCODE'], how='left').merge(sales_data, on=C['BARCODE'], how='left')

        final_cols = [C['BRAND'], C['BARCODE'], C['NAME'], C['SPEC'], C['STOCK'], C['PRICE'],
                      C['LAST_INBOUND_DATE'], C['REMARK'], C['TOTAL_REVENUE'], C['TOTAL_ORDERS'], C['TOTAL_SALES_QTY']] + week_labels
        final_data = final_data.reindex(columns=final_cols, fill_value=0)
        final_data[C['REMARK']] = final_data[C['REMARK']].fillna('')

        if self.progress_callback:
            self.progress_callback(75, "应用排序规则...")
        final_data = self._apply_sorting(final_data, sort_params)

        # V6.5 MODIFIED: Pass all necessary dataframes for weekly calculations
        if self.progress_callback:
            self.progress_callback(80, "创建报表文件...")
        if export_format.lower() == 'excel':
            report_path = self._create_and_save_excel(
                report_data=final_data,
                master_products=master_products,
                filtered_sales=filtered_sales,
                week_periods=week_periods,
                selected_brands=selected_brands,
                start_date=start_date,
                end_date=end_date,
                full_sales_df=sales_df,
                full_flow_df=flow_df,
                full_check_df=check_df
            )
        elif export_format.lower() == 'csv':
            report_path = self._create_and_save_csv(
                report_data=final_data,
                selected_brands=selected_brands,
                start_date=start_date,
                end_date=end_date
            )
        else:
            raise ValueError(f"不支持的导出格式: {export_format}")

        if self.progress_callback:
            self.progress_callback(100, "完成")
        if report_path:
            print(f"✅ 报表生成成功: {report_path}")
            return True, report_path
        return False, None

    def _get_week_periods(self, start_date, end_date):
        periods = []
        current_start = start_date
        while current_start <= end_date:
            week_end_day = current_start + timedelta(days=6 - current_start.weekday())
            current_end = min(week_end_day, end_date)
            periods.append((current_start, current_end))
            current_start = current_end + timedelta(days=1)
        return periods

    def _apply_sorting(self, data, sort_params):
        if data.empty:
            return data
        
        sort_columns = [p['field'] for p in sort_params]
        sort_ascending = [p['order'] == "升序" for p in sort_params]

        valid_cols = [col for col in sort_columns if col in data.columns]
        if valid_cols:
            data = data.sort_values(by=valid_cols, ascending=sort_ascending[:len(valid_cols)])

        return data.reset_index(drop=True)

    # V6.5 MODIFIED: Added full dataframes to the signature
    def _create_and_save_excel(self, report_data, master_products, filtered_sales, week_periods, selected_brands, start_date, end_date, full_sales_df, full_flow_df, full_check_df):
        wb = Workbook()
        ws = wb.active
        sheet_title = f"{start_date.strftime('%y%m%d')}-{end_date.strftime('%y%m%d')}总销售"
        ws.title = sheet_title[:31]

        styles = self._define_styles()
        week_labels = [f"{start.month}.{start.day}-{end.month}.{end.day}" for start, end in week_periods]
        if self.progress_callback:
            self.progress_callback(85, "写入总销售表...")
        self._write_sheet_data(ws, "总销售表", report_data, styles, week_labels)

        # V6.5 MODIFIED: Pass the full dataframes to the weekly sheet generator
        self._add_weekly_sheets(
            wb=wb,
            master_products=master_products,
            filtered_sales=filtered_sales,
            week_periods=week_periods,
            styles=styles,
            full_sales_df=full_sales_df,
            full_flow_df=full_flow_df,
            full_check_df=full_check_df
        )
        
        if self.progress_callback:
            self.progress_callback(95, "添加可视化图表...")
        # 添加可视化图表工作表
        self._add_visualization_sheet(wb, report_data, week_periods, styles)

        if self.progress_callback:
            self.progress_callback(98, "保存并增强兼容性...")
        # V8.0 MODIFIED: Use the enhanced save method
        return self._save_and_enhance_compatibility(wb, selected_brands, start_date, end_date)
    
    def _add_visualization_sheet(self, wb, report_data, week_periods, styles):
        """添加可视化图表工作表"""
        if report_data.empty:
            return
            
        C = Config.STD_COLS
        ws_chart = wb.create_sheet(title="可视化图表")
        
        # 添加标题
        title_font = Font(name='微软雅黑', size=14, bold=True)
        ws_chart['A1'] = "销售数据分析可视化"
        ws_chart['A1'].font = title_font
        
        # 1. 销量前10名商品柱状图
        top_products = report_data.nlargest(10, C['TOTAL_SALES_QTY'])
        if not top_products.empty:
            # 写入数据到A50以下
            data_start_row = 50
            ws_chart.cell(row=data_start_row, column=1, value="商品名称")
            ws_chart.cell(row=data_start_row, column=2, value="总销量")
            ws_chart.cell(row=data_start_row, column=3, value="总销售额")
            
            for i, (_, row) in enumerate(top_products.iterrows(), 1):
                ws_chart.cell(row=data_start_row+i, column=1, value=str(row[C['NAME']])[:15])  # 限制名称长度
                ws_chart.cell(row=data_start_row+i, column=2, value=row[C['TOTAL_SALES_QTY']])
                ws_chart.cell(row=data_start_row+i, column=3, value=row[C['TOTAL_REVENUE']])
            
            # 创建销量柱状图
            from openpyxl.chart import BarChart, Reference
            chart1 = BarChart()
            chart1.title = "销量前10名商品"
            chart1.x_axis.title = "商品"
            chart1.y_axis.title = "销量"
            
            # 数据范围
            categories = Reference(ws_chart, min_col=1, min_row=data_start_row+1, max_row=data_start_row+len(top_products))
            values = Reference(ws_chart, min_col=2, min_row=data_start_row, max_row=data_start_row+len(top_products))
            
            chart1.add_data(values, titles_from_data=True)
            chart1.set_categories(categories)
            # 设置图表大小为9x15cm
            chart1.width = 15
            chart1.height = 9
            
            # 将图表放置在A2格，以便露出标题
            ws_chart.add_chart(chart1, "A2")
            
            # 创建销售额柱状图
            chart2 = BarChart()
            chart2.title = "销售额前10名商品"
            chart2.x_axis.title = "商品"
            chart2.y_axis.title = "销售额"
            
            values2 = Reference(ws_chart, min_col=3, min_row=data_start_row, max_row=data_start_row+len(top_products))
            chart2.add_data(values2, titles_from_data=True)
            chart2.set_categories(categories)
            # 设置图表大小为9x15cm
            chart2.width = 15
            chart2.height = 9
            
            # 将图表放置在I2格，以便露出标题
            ws_chart.add_chart(chart2, "I2")
        
        # 2. 周度销量趋势折线图
        if week_periods:
            # 将周度数据也放在A50以下，但与上面的数据分开
            trend_start_row = 50 + 15  # 在销量数据之后
            ws_chart.cell(row=trend_start_row, column=1, value="周度期间")
            ws_chart.cell(row=trend_start_row, column=2, value="总销量")
            
            # 计算每周总销量
            week_labels = [f"{start.month}.{start.day}-{end.month}.{end.day}" for start, end in week_periods]
            weekly_totals = []
            for label in week_labels:
                if label in report_data.columns:
                    total = report_data[label].sum()
                    weekly_totals.append(total)
                else:
                    weekly_totals.append(0)
            
            # 写入周度数据
            for i, (label, total) in enumerate(zip(week_labels, weekly_totals), 1):
                ws_chart.cell(row=trend_start_row+i, column=1, value=label)
                ws_chart.cell(row=trend_start_row+i, column=2, value=total)
            
            # 创建折线图
            from openpyxl.chart import LineChart
            line_chart = LineChart()
            line_chart.title = "周度销量趋势"
            line_chart.x_axis.title = "周度期间"
            line_chart.y_axis.title = "销量"
            
            # 数据范围
            trend_categories = Reference(ws_chart, min_col=1, min_row=trend_start_row+1, max_row=trend_start_row+len(week_labels))
            trend_values = Reference(ws_chart, min_col=2, min_row=trend_start_row, max_row=trend_start_row+len(week_labels))
            
            line_chart.add_data(trend_values, titles_from_data=True)
            line_chart.set_categories(trend_categories)
            # 设置图表大小为9x15cm
            line_chart.width = 15
            line_chart.height = 9
            
            # 将图表放置在A22格
            ws_chart.add_chart(line_chart, "A22")
        
        # 3. 断货提醒 - 优化逻辑
        # 从Q2单元格开始显示
        stockout_start_row = 2  # 第2行
        stockout_start_col = 17  # Q列是第17列
        
        ws_chart.cell(row=stockout_start_row, column=stockout_start_col, value="断货提醒")
        ws_chart.cell(row=stockout_start_row, column=stockout_start_col).font = Font(name='微软雅黑', size=12, bold=True)
        
        # 筛选断货商品的优化逻辑：
        # 1. 库存为0且总销量>0的商品
        # 2. 库存极少（1-2个）但销量较高的商品（按周期内平均每周销量判断）
        stockout_candidates = report_data[report_data[C['STOCK']] <= 2].copy()
        
        if not stockout_candidates.empty and week_periods:
            # 计算每周平均销量
            num_weeks = len(week_periods)
            stockout_candidates['avg_weekly_sales'] = stockout_candidates[C['TOTAL_SALES_QTY']] / num_weeks
            
            # 筛选断货提醒商品：
            # 1. 库存为0且有销售记录的商品
            # 2. 库存极低（1-2个）但平均每周销量大于库存量的商品（表示可能很快售罄）
            stockout_products = stockout_candidates[
                (stockout_candidates[C['STOCK']] == 0) & (stockout_candidates[C['TOTAL_SALES_QTY']] > 0) |
                (stockout_candidates[C['STOCK']].isin([1, 2])) & (stockout_candidates['avg_weekly_sales'] > stockout_candidates[C['STOCK']])
            ].nlargest(20, C['TOTAL_SALES_QTY'])  # 取前20个
            
            if not stockout_products.empty:
                # 显示表头
                headers = [C['NAME'], C['TOTAL_SALES_QTY'], C['TOTAL_REVENUE'], C['STOCK'], '平均周销量']
                header_names = ['商品名称', '总销量', '总销售额', '库存量', '平均周销量']
                
                for col_offset, header_name in enumerate(header_names):
                    ws_chart.cell(row=stockout_start_row+1, column=stockout_start_col+col_offset, value=header_name)
                    ws_chart.cell(row=stockout_start_row+1, column=stockout_start_col+col_offset).font = Font(name='微软雅黑', size=10, bold=True)
                
                # 填充数据
                for row_offset, (_, product) in enumerate(stockout_products.iterrows(), stockout_start_row+2):
                    ws_chart.cell(row=row_offset, column=stockout_start_col, value=str(product[C['NAME']])[:20])
                    ws_chart.cell(row=row_offset, column=stockout_start_col+1, value=product[C['TOTAL_SALES_QTY']])
                    ws_chart.cell(row=row_offset, column=stockout_start_col+2, value=product[C['TOTAL_REVENUE']])
                    ws_chart.cell(row=row_offset, column=stockout_start_col+3, value=product[C['STOCK']])
                    ws_chart.cell(row=row_offset, column=stockout_start_col+4, value=round(product['avg_weekly_sales'], 2))
                    
                    # 对库存为0或极低的单元格添加特殊格式
                    if product[C['STOCK']] <= 2:
                        ws_chart.cell(row=row_offset, column=stockout_start_col+3).font = Font(color="FF0000", bold=True)  # 红色加粗显示

    def _create_and_save_csv(self, report_data, selected_brands, start_date, end_date):
        """创建并保存CSV格式报表"""
        try:
            brand_name = re.sub(r'[<>:"/\\|?*]', '', selected_brands[0])
            if len(selected_brands) > 1:
                brand_name += "等"
            ts = datetime.now().strftime('%H%M%S')
            filename = f"{brand_name}_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%m%d')}_{ts}.csv"
            report_path = Config.get_report_path(filename)
            
            # 保存CSV文件
            report_data.to_csv(report_path, index=False, encoding='utf-8-sig')
            return report_path
        except PermissionError:
            messagebox.showerror("文件保存失败", f"请关闭已打开的CSV文件 '{os.path.basename(filename)}' 后重试。")
            return None
        except Exception as e:
            messagebox.showerror("文件保存失败", f"无法保存CSV文件:\n{e}")
            return None

    def _define_styles(self):
        b = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        return {
            'header': {
                'font': Font(name='微软雅黑', size=11, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': b
            },
            'normal': {
                'font': Font(name='微软雅黑', size=10),
                'border': b,
                'alignment': Alignment(shrink_to_fit=True, vertical='center')
            },
            'totals': {
                'font': Font(name='微软雅黑', size=10, bold=True),
                'border': b
            },
            'remark_special': {
                'font': Font(name='微软雅黑', size=10, color='FF6600', bold=True),
                'fill': PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')
            },
            'low_stock': {
                'font': Font(name='微软雅黑', size=10, color='FF0000'),
                'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            }
        }

    def _write_sheet_data(self, ws, table_name, report_data, styles, week_labels=None):
        if week_labels is None:
            week_labels = []

        headers = list(report_data.columns)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = styles['header']['font']
            cell.fill = styles['header']['fill']
            cell.alignment = styles['header']['alignment']
            cell.border = styles['header']['border']

        C = Config.STD_COLS
        formats = {
            C['STOCK']: EXCEL_FORMATS['integer'],
            C['PRICE']: EXCEL_FORMATS['currency'],
            C['TOTAL_REVENUE']: EXCEL_FORMATS['currency'],
            C['TOTAL_ORDERS']: EXCEL_FORMATS['integer'],
            C['TOTAL_SALES_QTY']: EXCEL_FORMATS['integer'],
            '工作日销量': EXCEL_FORMATS['integer'],
            '周末销量': EXCEL_FORMATS['integer'],
            **{label: EXCEL_FORMATS['integer'] for label in week_labels}
        }

        for r_idx, row in enumerate(report_data.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                col_name = headers[c_idx - 1]

                if col_name == C['SPEC'] and (value == 0 or str(value) == '0' or pd.isna(value)):
                    cell.value = None
                else:
                    cell.value = value

                cell.font = styles['normal']['font']
                cell.border = styles['normal']['border']
                cell.alignment = styles['normal']['alignment']
                
                if col_name == C['REMARK'] and value == '商品可能已退库':
                    cell.font = styles['remark_special']['font']
                    cell.fill = styles['remark_special']['fill']
                
                is_low_stock = row[headers.index(C['STOCK'])] < 2
                is_not_returned = row[headers.index(C['REMARK'])] != '商品可能已退库'
                if col_name == C['STOCK'] and is_low_stock and is_not_returned:
                     cell.font = styles['low_stock']['font']
                     cell.fill = styles['low_stock']['fill']

                if col_name in formats and isinstance(value, (int, float, np.number)):
                    if value == 0 and formats[col_name] == EXCEL_FORMATS['integer'] and col_name != C['STOCK']:
                        cell.value = None
                    else:
                        cell.number_format = formats[col_name]
        
        # V8.0 MODIFIED: Adopted the summation logic from v7.1
        if not report_data.empty:
            totals_row_index = len(report_data) + 2
            sum_cols = [
                C['STOCK'], C['TOTAL_REVENUE'], C['TOTAL_ORDERS'],
                C['TOTAL_SALES_QTY'], '工作日销量', '周末销量'
            ] + week_labels
            
            for c_idx, col_name in enumerate(headers, 1):
                total_cell = ws.cell(row=totals_row_index, column=c_idx)

                total_cell.font = styles['totals']['font']
                total_cell.border = styles['totals']['border']

                if col_name == C['NAME']:
                    total_cell.value = f'=SUBTOTAL(103,[{C["NAME"]}])&"个SKU"'
                elif col_name in sum_cols:
                    total_cell.value = f'=SUBTOTAL(109,[{col_name}])'
                    if col_name in formats:
                        total_cell.number_format = formats[col_name]
        
        table_columns = [
            TableColumn(id=i + 1, name=col_name) for i, col_name in enumerate(headers)
        ]

        table_ref = f"A1:{get_column_letter(len(headers))}{len(report_data) + 2}"
        table = Table(displayName=table_name.replace(" ", ""), ref=table_ref, tableColumns=table_columns, totalsRowCount=1, totalsRowShown=True)
        style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        for column_cells in ws.columns:
            header_cell = column_cells[0]
            header_text = header_cell.value
            column_letter = header_cell.column_letter
            
            if header_text in week_labels:
                ws.column_dimensions[column_letter].width = 10
                continue
            if header_text in ['工作日销量', '周末销量']:
                ws.column_dimensions[column_letter].width = 11
                continue
            if header_text == C['REMARK']:
                ws.column_dimensions[column_letter].width = 13
                continue
            if header_text == C['TOTAL_REVENUE']:
                ws.column_dimensions[column_letter].width = 12
                continue

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.row == len(report_data) + 2:
                        continue
                    if cell.value:
                        cell_len = len(str(cell.value).encode('gbk', 'ignore')) if any('\u4e00' <= char <= '\u9fff' for char in str(cell.value)) else len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 8), 50) 
            
            if header_text == C['NAME']:
                adjusted_width = max(adjusted_width, 40)
                
            ws.column_dimensions[column_letter].width = adjusted_width

    # V8.0 MODIFIED: Replaced _save_workbook with _save_and_enhance_compatibility from v7.1
    def _save_and_enhance_compatibility(self, wb, selected_brands, start_date, end_date):
        try:
            brand_name = re.sub(r'[<>:"/\\|?*]', '', selected_brands[0])
            if len(selected_brands) > 1:
                brand_name += "等"
            ts = datetime.now().strftime('%H%M%S')
            filename = f"{brand_name}_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%m%d')}_{ts}.xlsx"
            report_path = Config.get_report_path(filename)

            # 步骤1: 先由 openpyxl 保存文件
            wb.save(report_path)

            # 步骤2: 调用Excel/WPS COM组件重新打开并保存，以修复移动端兼容性问题
            abs_report_path = os.path.abspath(report_path)
            com_ids = ['ket.Application', 'et.Application', 'wps.application', 'Kingsoft.Application', 'Excel.Application']  # 优先用ket.Application, 再尝试其他WPS ID, 最后是Excel
            re_saved_successfully = False

            for com_id in com_ids:
                excel = None
                try:
                    import win32com.client as win32
                    import traceback
                    print(f"🔄 尝试使用 {com_id} 引擎重新保存...")
                    excel = win32.DispatchEx(com_id)
                    excel.Visible = False
                    excel.DisplayAlerts = False

                    workbook = excel.Workbooks.Open(abs_report_path)
                    workbook.Save()
                    workbook.Close(SaveChanges=False)
                    excel.Quit()

                    print(f"✅ 使用 {com_id} 兼容性保存成功！")
                    re_saved_successfully = True
                    break  # Exit loop on success
                except Exception as e:
                    if excel:
                        excel.Quit()
                    print(f"  -> {com_id} 不可用或失败。")
                    # print(f"   [调试日志] 错误: {e}") # 移除详细的traceback避免干扰普通用户
                    continue  # Try next COM ID

            if not re_saved_successfully:
                print("⚠️ 未能通过任何Excel/WPS引擎重新保存。报表已生成，但可能在移动端显示不佳。")
                print("   请确保已安装 Microsoft Excel 或 WPS Office，并已安装 pywin32 库。")

            return report_path
        except PermissionError:
            messagebox.showerror("文件保存失败", f"请关闭已打开的Excel文件 '{os.path.basename(filename)}' 后重试。")
            return None
        except Exception as e:
            messagebox.showerror("文件保存失败", f"无法保存Excel文件:\n{e}")
            return None

    # V6.5 MODIFIED: Added full dataframes to signature for weekly calculation
    def _add_weekly_sheets(self, wb, master_products, filtered_sales, week_periods, styles, full_sales_df, full_flow_df, full_check_df):
        print("📅 正在生成周度报表(v6.5 独立库存模式)...")
        C = Config.STD_COLS

        if not week_periods:
            print("ℹ️ 在选定范围内未找到任何期间。")
            return

        num_weeks = len(week_periods)
        for i, (week_start, week_end) in enumerate(week_periods):
            if self.progress_callback and num_weeks > 0:
                progress = 85 + int(((i + 1) / num_weeks) * 10) # 85% to 95%
                self.progress_callback(progress, f"正在生成周度报表: {i+1}/{num_weeks}")

            week_start_inclusive = datetime.combine(week_start, datetime.min.time())
            week_end_inclusive = datetime.combine(week_end, datetime.max.time())

            weekly_sales = filtered_sales[
                (filtered_sales[C['SALES_TIME']] >= week_start_inclusive) &
                (filtered_sales[C['SALES_TIME']] <= week_end_inclusive)
            ].copy()

            if weekly_sales.empty:
                continue

            weekly_summary = weekly_sales.groupby(C['BARCODE']).agg(
                total_revenue=(C['REVENUE'], 'sum'),
                total_orders=(C['ORDER_ID'], 'nunique'),
                total_sales_qty=(C['SALES_QTY'], 'sum')
            ).rename(columns={
                'total_revenue': C['TOTAL_REVENUE'],
                'total_orders': C['TOTAL_ORDERS'],
                'total_sales_qty': C['TOTAL_SALES_QTY']
            })

            weekly_sales['weekday'] = weekly_sales[C['SALES_TIME']].dt.weekday
            weekday_sales_series = weekly_sales[weekly_sales['weekday'] < 5].groupby(C['BARCODE'])[C['SALES_QTY']].sum().rename('工作日销量')
            weekend_sales_series = weekly_sales[weekly_sales['weekday'] >= 5].groupby(C['BARCODE'])[C['SALES_QTY']].sum().rename('周末销量')
            
            weekly_sales_summary = pd.concat([weekly_summary, weekday_sales_series, weekend_sales_series], axis=1).fillna(0).reset_index()

            sold_barcodes = weekly_sales_summary[C['BARCODE']].unique()
            if len(sold_barcodes) == 0:
                continue
            
            # V6.5 MODIFIED: Calculate inventory specifically for this week's end date
            weekly_inventory_data = self.inventory_calc.calculate_inventory(
                product_barcodes=sold_barcodes,
                flow_df=full_flow_df,
                check_df=full_check_df,
                sales_df=full_sales_df,
                end_date=week_end  # Pass the end date of the current week
            )

            weekly_report_data = master_products[master_products[C['BARCODE']].isin(sold_barcodes)].copy()
            # V6.5 MODIFIED: Merge the newly calculated weekly inventory
            weekly_report_data = weekly_report_data.merge(weekly_inventory_data, on=C['BARCODE'], how='left')
            weekly_report_data = weekly_report_data.merge(weekly_sales_summary, on=C['BARCODE'], how='left')

            weekly_cols = [
                C['BRAND'], C['BARCODE'], C['NAME'], C['SPEC'], C['STOCK'], C['PRICE'],
                C['LAST_INBOUND_DATE'], C['REMARK'], C['TOTAL_REVENUE'], C['TOTAL_ORDERS'],
                C['TOTAL_SALES_QTY'], '工作日销量', '周末销量'
            ]
            weekly_report_data = weekly_report_data.reindex(columns=weekly_cols, fill_value=0)
            weekly_report_data[C['REMARK']] = weekly_report_data[C['REMARK']].fillna('')

            weekly_report_data = self._apply_sorting(weekly_report_data, [{'field': C['REMARK'], 'order': '升序'}, {'field': C['TOTAL_SALES_QTY'], 'order': '降序'}])

            sheet_name = f"W_{week_start.strftime('%m%d')}-{week_end.strftime('%m%d')}"
            ws_week = wb.create_sheet(title=sheet_name[:31])
            print(f"  - 创建工作表: {sheet_name}")

            self._write_sheet_data(ws_week, f"WeekTable{i}", weekly_report_data, styles)


# ==================== 主GUI界面 ====================
class SupplierReportGUI:
    def __init__(self, root):
        self.root = root
        # V8.0 MODIFIED: Update window title
        self.root.title("供应商报表生成器 - v8.0")
        self.root.geometry("900x800")
        self.root.minsize(850, 750)
        style = ttk.Style(self.root)
        try:
            style.theme_use('vista')
        except tk.TclError:
            print("Vista theme not available, using default.")
        
        style.configure("Accent.TButton", foreground="black", background="#0078D7")
        
        Config.ensure_folders()
        self.data_processor = DataProcessor()
        self.product_manager = ProductManager(self.data_processor)
        self.inventory_calc = InventoryCalculator(self.data_processor)
        self.sales_analyzer = SalesAnalyzer(self.data_processor)
        self.data_status_manager = DataStatusManager()
        self.data_frames, self.product_file_path, self.all_brands, self.brand_vars = {}, None, [], {}
        self.reference_date = datetime.now()
        self.end_date = self.reference_date
        self.start_date = self.reference_date - timedelta(days=29)
        self.export_format = tk.StringVar(value="excel")  # 默认导出格式
        self.mouse_in_brand_area = False  # 跟踪鼠标是否在品牌区域
        self.setup_ui()
        self.last_selected_brand_count = 0
        self.load_data_with_progress()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.grid_rowconfigure(2, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        title = ttk.Label(main_frame, text="供应商销售报表生成器", font=("微软雅黑", 18, "bold"))
        title.grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="w")

        self._create_data_status_area(main_frame).grid(row=1, column=0, sticky="ew", pady=10)
        self._create_brand_selection_area(main_frame).grid(row=2, column=0, sticky="nsew", pady=10)

        settings_frame = ttk.Frame(main_frame)
        settings_frame.grid(row=3, column=0, sticky="ew", pady=10)
        settings_frame.grid_columnconfigure(0, weight=1)
        settings_frame.grid_columnconfigure(1, weight=1)

        self._create_time_selection_area(settings_frame).grid(row=0, column=0, sticky="ewns", padx=(0, 5))
        self._create_sort_settings_area(settings_frame).grid(row=0, column=1, sticky="ewns", padx=(5, 0))

        self._create_action_buttons(main_frame).grid(row=4, column=0, sticky="ew", pady=15)

        self.status_var = tk.StringVar(value="准备就绪")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=5)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def _create_data_status_area(self, parent):
        status_frame = ttk.LabelFrame(parent, text="数据源状态", padding="10")
        self.status_text = tk.Text(status_frame, height=9, width=80, font=("微软雅黑", 9), bg="#f0f0f0", relief=tk.FLAT, wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True)
        self.update_data_status_display("正在初始化...")
        return status_frame

    def _create_brand_selection_area(self, parent):
        brand_frame = ttk.LabelFrame(parent, text="品牌选择", padding="10")
        brand_frame.grid_rowconfigure(1, weight=1)
        brand_frame.grid_columnconfigure(0, weight=1)

        search_frame = ttk.Frame(brand_frame)
        search_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        ttk.Label(search_frame, text="搜索品牌:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_brands())
        ttk.Entry(search_frame, textvariable=self.search_var).pack(side=tk.LEFT, padx=(5, 0), fill=tk.X, expand=True)

        canvas_frame = ttk.Frame(brand_frame)
        canvas_frame.grid(row=1, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(canvas_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # 绑定鼠标进入和离开事件
        self.canvas.bind("<Enter>", self._on_mouse_enter_brand_area)
        self.canvas.bind("<Leave>", self._on_mouse_leave_brand_area)
        # 只在品牌区域响应滚轮事件
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        
        btn_frame = ttk.Frame(brand_frame)
        btn_frame.grid(row=2, column=0, sticky="ew", pady=(5, 0))
        ttk.Button(btn_frame, text="全选", command=self.select_all_brands).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_frame, text="清空", command=self.clear_brand_selection).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="反选", command=self.invert_brand_selection).pack(side=tk.LEFT, padx=5)

        return brand_frame

    def _on_mouse_enter_brand_area(self, event):
        """鼠标进入品牌区域"""
        self.mouse_in_brand_area = True

    def _on_mouse_leave_brand_area(self, event):
        """鼠标离开品牌区域"""
        self.mouse_in_brand_area = False

    def _on_mousewheel(self, event):
        """只在品牌区域处理滚轮事件"""
        if self.mouse_in_brand_area:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _create_time_selection_area(self, parent):
        time_frame = ttk.LabelFrame(parent, text="时间范围设置", padding="10")
        self.start_date_var = tk.StringVar(value=self.start_date.strftime("%Y-%m-%d"))
        self.end_date_var = tk.StringVar(value=self.end_date.strftime("%Y-%m-%d"))

        ttk.Label(time_frame, text="选择时间范围:").grid(row=0, column=0, sticky="w")
        date_entry_frame = ttk.Frame(time_frame)
        date_entry_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=5)
        ttk.Entry(date_entry_frame, textvariable=self.start_date_var, width=12, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Label(date_entry_frame, text=" 至 ").pack(side=tk.LEFT, padx=5)
        ttk.Entry(date_entry_frame, textvariable=self.end_date_var, width=12, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(time_frame, text="选择...", command=self.select_date_range).grid(row=1, column=2, padx=(5, 0), sticky="ew")
        
        # 导出格式选择
        ttk.Label(time_frame, text="导出格式:").grid(row=2, column=0, sticky="w", pady=(10, 0))
        format_frame = ttk.Frame(time_frame)
        format_frame.grid(row=3, column=0, columnspan=2, sticky="w", pady=5)
        ttk.Radiobutton(format_frame, text="Excel (.xlsx)", variable=self.export_format, value="excel").pack(side=tk.LEFT)
        ttk.Radiobutton(format_frame, text="CSV (.csv)", variable=self.export_format, value="csv").pack(side=tk.LEFT, padx=(10, 0))
        
        return time_frame

    def _create_sort_settings_area(self, parent):
        sort_frame = ttk.LabelFrame(parent, text="排序设置", padding="10")
        sort_frame.grid_columnconfigure(0, weight=1)
        self.sort_rules = []
        self.sort_frame_container = ttk.Frame(sort_frame)
        self.sort_frame_container.grid(row=0, column=0, sticky="ew")

        btn_frame = ttk.Frame(sort_frame)
        btn_frame.grid(row=1, column=0, sticky="w", pady=(10, 0))
        ttk.Button(btn_frame, text="+ 添加排序", command=self.add_sort_rule).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="- 删除排序", command=self.remove_sort_rule).pack(side=tk.LEFT, padx=5)

        self.add_sort_rule(Config.STD_COLS['REMARK'], "升序")
        self.add_sort_rule(Config.STD_COLS['TOTAL_SALES_QTY'], "降序")

        return sort_frame

    def _create_action_buttons(self, parent):
        btn_frame = ttk.Frame(parent)
        ttk.Button(btn_frame, text="🚀 生成报表", command=self.generate_report, padding=10, style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="退出", command=self.root.quit).pack(side=tk.LEFT)
        return btn_frame

    def load_data_with_progress(self):
        progress_dialog = ProgressDialog(self.root, "正在加载数据...")

        def load_data_thread():
            try:
                progress_dialog.update_progress(10, "加载和预处理数据...")
                self.data_frames, self.product_file_path = self.product_manager.load_and_prep_data()
                sales_df = self.data_frames.get('sales')
                if sales_df is None or sales_df.empty:
                    self.root.after(0, lambda: [messagebox.showerror("严重错误", "销售数据 (sales_data.xlsx) 未找到或为空。" ), progress_dialog.destroy()])
                    return

                progress_dialog.update_progress(60, "分析品牌信息...")
                self.all_brands = self.product_manager.get_all_brands(self.data_frames['product'], sales_df)
                if not self.all_brands:
                    self.root.after(0, lambda: [messagebox.showerror("错误", "无法找到品牌信息。" ), progress_dialog.destroy()])
                    return

                progress_dialog.update_progress(80, "更新数据状态...")
                self.data_status_manager.update_all_statuses(self.data_frames, self.product_file_path)
                last_sale = sales_df[Config.STD_COLS['SALES_TIME']].max()
                self.reference_date = datetime(last_sale.year, last_sale.month, last_sale.day)
                self.end_date, self.start_date = self.reference_date, self.reference_date - timedelta(days=29)

                progress_dialog.update_progress(100, "加载完成！")
                self.root.after(0, lambda: [self.finalize_data_loading(), progress_dialog.destroy()])
            except Exception as e:
                self.root.after(0, lambda: [messagebox.showerror("加载错误", f"数据加载时出错: {e}"), progress_dialog.destroy()])
                import traceback
                traceback.print_exc()

        threading.Thread(target=load_data_thread, daemon=True).start()

    def finalize_data_loading(self):
        self.start_date_var.set(self.start_date.strftime("%Y-%m-%d"))
        self.end_date_var.set(self.end_date.strftime("%Y-%m-%d"))
        self.create_brand_checkboxes()
        self.update_data_status_display(self.data_status_manager.get_status_display_text())
        self.status_var.set(f"数据加载完成，共找到 {len(self.all_brands)} 个品牌。")
        
        # 执行数据质量检查
        quality_issues = DataQualityChecker.check_data_quality(self.data_frames)
        if quality_issues:
            issue_text = "\n".join([f"⚠️ {issue}" for issue in quality_issues])
            self.update_data_status_display(
                self.data_status_manager.get_status_display_text() + 
                f"\n\n数据质量检查发现问题:\n{issue_text}"
            )

    def update_data_status_display(self, text):
        self.status_text.config(state=tk.NORMAL)
        self.status_text.delete(1.0, tk.END)
        self.status_text.insert(1.0, text)
        self.status_text.config(state=tk.DISABLED)

    def create_brand_checkboxes(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        search_text = self.search_var.get().lower()
        brands_to_show = [b for b in self.all_brands if search_text in b.lower()] if search_text else self.all_brands
        self.brand_vars = {brand: self.brand_vars.get(brand, tk.BooleanVar()) for brand in self.all_brands}
        for brand in brands_to_show:
            cb = ttk.Checkbutton(self.scrollable_frame, text=brand, variable=self.brand_vars[brand])
            cb.pack(anchor=tk.W, padx=5, pady=2)
            self.brand_vars[brand].trace_add('write', self.on_brand_selection_change)
        
        # 根据显示项数量决定是否启用滚轮
        # 当匹配项>=5时允许滚动，并设置合适的滚动范围
        self.root.unbind_all("<MouseWheel>")  # 解除全局绑定
        if len(brands_to_show) >= 5:
            # 计算合适的滚动区域，避免滚动过多
            item_height = 25  # 估算每个品牌项的高度
            total_height = len(brands_to_show) * item_height
            visible_height = 10 * item_height  # 假设可显示10个项目
            if total_height > visible_height:
                # 设置滚动区域
                self.canvas.configure(scrollregion=(0, 0, 0, total_height))
        else:
            # 禁用滚动
            self.canvas.configure(scrollregion=(0, 0, 0, 0))

    def filter_brands(self):
        self.create_brand_checkboxes()

    def select_all_brands(self):
        for var in self._get_visible_brand_vars():
            var.set(True)

    def clear_brand_selection(self):
        for var in self._get_visible_brand_vars():
            var.set(False)

    def invert_brand_selection(self):
        for var in self._get_visible_brand_vars():
            var.set(not var.get())

    def _get_visible_brand_vars(self):
        search_text = self.search_var.get().lower()
        if not search_text:
            return self.brand_vars.values()
        return [var for brand, var in self.brand_vars.items() if search_text in brand.lower()]

    def get_selected_brands(self):
        return [brand for brand, var in self.brand_vars.items() if var.get()]

    def on_brand_selection_change(self, *args):
        current_count = len(self.get_selected_brands())
        
        is_multi_now = current_count > 1
        was_multi_before = self.last_selected_brand_count > 1

        if is_multi_now != was_multi_before:
            self.update_sort_rules_for_brand_count(current_count)

        self.last_selected_brand_count = current_count

    def update_sort_rules_for_brand_count(self, count):
        while self.sort_rules:
            self.remove_sort_rule()

        C = Config.STD_COLS
        if count > 1:
            print("🔄 检测到多品牌选择，切换排序规则...")
            self.add_sort_rule(C['REMARK'], "升序")
            self.add_sort_rule(C['BRAND'], "升序")
            self.add_sort_rule(C['TOTAL_SALES_QTY'], "降序")
        else:
            print("🔄 检测到单品牌选择，切换排序规则...")
            self.add_sort_rule(C['REMARK'], "升序")
            self.add_sort_rule(C['TOTAL_SALES_QTY'], "降序")

    def add_sort_rule(self, default_field=Config.STD_COLS['BRAND'], default_order="升序"):
        if len(self.sort_rules) >= 5:
            return
        rule_frame = ttk.Frame(self.sort_frame_container)
        rule_frame.pack(fill=tk.X, pady=2)

        field_var = tk.StringVar(value=default_field)
        order_var = tk.StringVar(value=default_order)

        C = Config.STD_COLS
        sortable_cols = [
            C['BRAND'], C['BARCODE'], C['NAME'], C['SPEC'], C['STOCK'], C['PRICE'],
            C['LAST_INBOUND_DATE'], C['REMARK'], C['TOTAL_REVENUE'], C['TOTAL_ORDERS'], C['TOTAL_SALES_QTY']
        ]

        ttk.Label(rule_frame, text=f"排序 {len(self.sort_rules) + 1}:").pack(side=tk.LEFT, padx=(0, 5))
        field_combo = ttk.Combobox(rule_frame, textvariable=field_var, values=sortable_cols, width=12, state="readonly")
        field_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        order_combo = ttk.Combobox(rule_frame, textvariable=order_var, values=["升序", "降序"], width=8, state="readonly")
        order_combo.pack(side=tk.LEFT, padx=5)

        self.sort_rules.append({'frame': rule_frame, 'field_var': field_var, 'order_var': order_var})

    def remove_sort_rule(self):
        if len(self.sort_rules) > 0:
            rule = self.sort_rules.pop()
            rule['frame'].destroy()

    def get_sort_params(self):
        return [{'field': r['field_var'].get(), 'order': r['order_var'].get()} for r in self.sort_rules]

    def select_date_range(self):
        def on_date_selected(start, end):
            self.start_date, self.end_date = start, end
            self.start_date_var.set(start.strftime("%Y-%m-%d"))
            self.end_date_var.set(end.strftime("%Y-%m-%d"))

        DatePickerWidget(self.root, self.start_date, self.end_date, on_date_selected, self.reference_date)

    def generate_report(self):
        selected_brands = self.get_selected_brands()
        if not selected_brands:
            messagebox.showwarning("警告", "请至少选择一个品牌。" )
            return

        progress_dialog = ProgressDialog(self.root, "正在生成报表...")
        self.status_var.set("正在生成报表，请稍候...")
        self.root.update()

        def report_thread():
            try:
                start_date, end_date = self.start_date, self.end_date
                export_format = self.export_format.get()
                report_generator = ReportGenerator(self.data_processor, self.inventory_calc, self.sales_analyzer, self.product_manager)
                
                report_generator.set_progress_callback(progress_dialog.update_progress)
                
                success, report_path = report_generator.generate_report(
                    self.data_frames, selected_brands, start_date, end_date, self.get_sort_params(), export_format)
                
                self.root.after(0, progress_dialog.destroy)

                if success and report_path:
                    self.status_var.set(f"报表生成成功！已保存至 {os.path.basename(report_path)}")
                    self.root.after(0, lambda: self._show_success_dialog(report_path))
                else:
                    self.status_var.set("报表生成失败，请检查数据文件和设置。")
            except Exception as e:
                self.root.after(0, progress_dialog.destroy)
                self.status_var.set(f"发生严重错误: {e}")
                self.root.after(0, lambda: messagebox.showerror("严重错误", f"生成报表时发生意外错误:\n{e}"))
                import traceback
                traceback.print_exc()
        
        threading.Thread(target=report_thread, daemon=True).start()

    def _show_success_dialog(self, report_path):
        buttons = [
            ("关闭", None),
            ("打开所在文件夹", lambda: open_file_or_folder(os.path.dirname(report_path))),
            ("打开文件", lambda: open_file_or_folder(report_path))
        ]
        dialog = CustomMessageBox(self.root, "生成成功", f"报表已成功生成！\n\n文件路径:\n{report_path}", buttons)
        self.root.wait_window(dialog)

# ==================== 进度条和对话框组件 ====================
class ProgressDialog(tk.Toplevel):
    def __init__(self, parent, title="正在加载数据..."):
        super().__init__(parent)
        self.title(title)
        self.geometry("550x200")
        self.resizable(False, False)
        self.transient(parent)
        self.center_window()
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="准备加载...")
        self.create_widgets()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="数据处理中，请稍候...", font=("微软雅黑", 12, "bold")).pack(pady=(0, 10))
        
        ttk.Label(main_frame, textvariable=self.status_var, font=("微软雅黑", 9)).pack(pady=(0, 10))

        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, padx=10, pady=5)

        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, mode='determinate', maximum=100)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

        self.percent_var = tk.StringVar(value="0%")
        self.percent_label = ttk.Label(progress_frame, textvariable=self.percent_var, font=("微软雅黑", 9, "bold"), width=5, anchor='e')
        self.percent_label.pack(side=tk.RIGHT, padx=(10, 0))

    def update_progress(self, value, status):
        target_value = int(value)
        current_value = int(self.progress_var.get())
        
        self.status_var.set(status)
        self.update_idletasks()

        if target_value <= current_value:
            if target_value == 100: # Ensure 100% is set
                self.progress_var.set(target_value)
                self.percent_var.set(f"{target_value}%")
                self.update_idletasks()
            return

        # Animate the progress
        for i in range(current_value + 1, target_value + 1):
            self.progress_var.set(i)
            self.percent_var.set(f"{i}%")
            self.update_idletasks()
            import time
            time.sleep(0.015)
        
    def destroy(self):
        super().destroy()

class CustomMessageBox(tk.Toplevel):
    def __init__(self, parent, title, message, buttons):
        super().__init__(parent)
        self.title(title)
        self.transient(parent)
        self.update_idletasks()
        p_x, p_y, p_w, p_h = parent.winfo_x(), parent.winfo_y(), parent.winfo_width(), parent.winfo_height()
        w, h = 450, 180
        x, y = p_x + (p_w // 2) - (w // 2), p_y + (p_h // 2) - (h // 2)
        self.geometry(f'{w}x{h}+{x}+{y}')
        self.resizable(False, False)
        self.grab_set()

        main_frame = ttk.Frame(self, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        msg_frame = ttk.Frame(main_frame)
        msg_frame.pack(fill=tk.X, expand=True, pady=(0, 20))
        try:
            ttk.Label(msg_frame, image="::tk::icons::information", padding=(0, 0, 10, 0)).pack(side=tk.LEFT, anchor=tk.N)
        except tk.TclError:
            ttk.Label(msg_frame, text="✅", font=("Segoe UI Emoji", 16)).pack(side=tk.LEFT, anchor=tk.N, padx=(0, 10))
        ttk.Label(msg_frame, text=message, wraplength=350, justify=tk.LEFT).pack(side=tk.LEFT, fill=tk.X, expand=True)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        for i, (text, cmd) in enumerate(reversed(buttons)):
            is_accent = (i == 0)
            btn = ttk.Button(btn_frame, text=text, command=lambda c=cmd: self._execute_and_close(c))
            if is_accent:
                 btn.config(style="Accent.TButton")
            btn.pack(side=tk.RIGHT, padx=(5, 0))
            
    def _execute_and_close(self, command):
        if command:
            command()
        self.destroy()

class DatePickerWidget(tk.Toplevel):
    def __init__(self, parent, initial_start_date=None, initial_end_date=None, callback=None, reference_date=None):
        super().__init__(parent)
        self.callback = callback
        self.reference_date = reference_date or datetime.now()
        self.start_date = initial_start_date or self.reference_date
        self.end_date = initial_end_date or self.reference_date
        self.current_display_date = self.start_date
        self.selecting_start = True

        self.title("选择时间范围")
        self.geometry("800x650")
        self.resizable(True, True)
        self.transient(parent)
        self.create_widgets()
        self.center_window()
        self.grab_set()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (self.winfo_width() // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

    def create_widgets(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        main_frame = ttk.Frame(self, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.grid_rowconfigure(1, weight=1)
        main_frame.grid_columnconfigure(0, weight=1)

        self._create_time_range_display(main_frame).grid(row=0, column=0, sticky="ew", pady=(0, 10))
        middle_frame = ttk.Frame(main_frame)
        middle_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        middle_frame.grid_rowconfigure(0, weight=1)
        middle_frame.grid_columnconfigure(0, weight=1)
        self._create_calendar_view(middle_frame).grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self._create_shortcuts_menu(middle_frame).grid(row=0, column=1, sticky="ns")
        self._create_bottom_buttons(main_frame).grid(row=2, column=0, sticky="ew")

    def _create_time_range_display(self, parent):
        frame = ttk.LabelFrame(parent, text="选择的时间范围", padding="10")
        self.status_label = ttk.Label(frame, text="请点击日历选择开始日期", font=("微软雅黑", 10), foreground="green")
        self.status_label.pack(fill=tk.X, pady=(0, 10))
        range_frame = ttk.Frame(frame)
        range_frame.pack(fill=tk.X)
        self.start_date_label = self._create_date_label(range_frame, "开始时间:", self.start_date)
        self.end_date_label = self._create_date_label(range_frame, "结束时间:", self.end_date)
        ttk.Button(range_frame, text="重新选择", command=self.reset_selection, width=10).pack(side=tk.LEFT, padx=(20, 0), anchor='s')
        return frame

    def _create_date_label(self, parent, text, date_obj):
        frame = ttk.Frame(parent)
        frame.pack(side=tk.LEFT, padx=(0, 20))
        ttk.Label(frame, text=text, font=("微软雅黑", 10, "bold")).pack()
        label = ttk.Label(frame, text=date_obj.strftime("%Y-%m-%d"), font=("微软雅黑", 12), foreground="blue")
        label.pack()
        return label

    def _create_calendar_view(self, parent):
        frame = ttk.Frame(parent)
        frame.grid_rowconfigure(2, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        nav_frame = ttk.Frame(frame)
        nav_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        nav_frame.columnconfigure(1, weight=1)
        ttk.Button(nav_frame, text="<", command=self.prev_month, width=3).grid(row=0, column=0, sticky="w")
        
        year_month_frame = ttk.Frame(nav_frame)
        year_month_frame.grid(row=0, column=1, sticky="ew")
        self.year_label = tk.Label(year_month_frame, font=("微软雅黑", 12, "bold"), cursor="hand2", fg="blue")
        self.year_label.pack(side=tk.LEFT, padx=(5,0))
        self.year_label.bind("<Button-1>", self.show_year_selector)
        self.month_label = ttk.Label(year_month_frame, font=("微软雅黑", 12, "bold"))
        self.month_label.pack(side=tk.LEFT, padx=(5, 0))
        
        ttk.Button(nav_frame, text=">", command=self.next_month, width=3).grid(row=0, column=2, sticky="e")

        week_frame = ttk.Frame(frame)
        week_frame.grid(row=1, column=0, sticky="ew", pady=(0, 5))
        for i, day in enumerate(["一", "二", "三", "四", "五", "六", "日"]):
            week_frame.columnconfigure(i, weight=1)
            ttk.Label(week_frame, text=day, width=6, anchor=tk.CENTER, font=("微软雅黑", 10, "bold")).grid(row=0, column=i, sticky="ew")

        self.calendar_frame = ttk.Frame(frame)
        self.calendar_frame.grid(row=2, column=0, sticky="nsew")
        for i in range(6):
            self.calendar_frame.grid_rowconfigure(i, weight=1)
        for i in range(7):
            self.calendar_frame.grid_columnconfigure(i, weight=1)

        self.day_buttons = [[tk.Button(self.calendar_frame, text="", width=6, height=2, font=("微软雅黑", 10), relief=tk.FLAT, bd=1) for _ in range(7)] for _ in range(6)]
        for w, row in enumerate(self.day_buttons):
            for d, btn in enumerate(row):
                btn.grid(row=w, column=d, padx=1, pady=1, sticky="nsew")

        self.update_calendar()
        return frame

    def _create_shortcuts_menu(self, parent):
        frame = ttk.Frame(parent)
        ttk.Label(frame, text="快捷选择", font=("微软雅黑", 11, "bold")).pack(pady=(0, 10))

        ref_frame = ttk.Frame(frame)
        ref_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(ref_frame, text="参考日期 (最后销售日):", font=("微软雅黑", 9)).pack()
        ttk.Label(ref_frame, text=self.reference_date.strftime("%Y-%m-%d"), font=("微软雅黑", 9, "bold"), foreground="green").pack()

        shortcuts = [
            ("今天", self.set_today), ("昨天", self.set_yesterday),
            ("本周", self.set_this_week), ("上周", self.set_last_week),
            ("本月", self.set_this_month), ("上月", self.set_last_month),
            ("最近7天", self.set_last_7_days), ("最近30天", self.set_last_30_days),
            ("本季度", self.set_this_quarter), ("上季度", self.set_last_quarter)
        ]
        for text, cmd in shortcuts:
            ttk.Button(frame, text=text, command=cmd, width=12).pack(pady=2, fill=tk.X)
        return frame

    def _create_bottom_buttons(self, parent):
        frame = ttk.Frame(parent)
        ttk.Button(frame, text="取消", command=self.destroy, width=10).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(frame, text="确定", command=self.confirm, width=10, style="Accent.TButton").pack(side=tk.RIGHT)
        ttk.Button(frame, text="重置", command=self.reset_to_reference, width=10).pack(side=tk.LEFT)
        return frame

    def update_calendar(self):
        self.year_label.config(text=f"{self.current_display_date.year}年")
        self.month_label.config(text=f"{self.current_display_date.month:02d}月")
        cal = calendar.monthcalendar(self.current_display_date.year, self.current_display_date.month)
        today = date.today()
        start_d, end_d = self.start_date.date(), self.end_date.date()

        for week_idx in range(6):
            if week_idx < len(cal):
                week = cal[week_idx]
                for day_idx, day in enumerate(week):
                    btn = self.day_buttons[week_idx][day_idx]
                    if day == 0:
                        btn.config(text="", state=tk.DISABLED, bg="SystemButtonFace", relief=tk.FLAT)
                        continue
                    
                    btn.config(text=str(day), state=tk.NORMAL)
                    current_d = date(self.current_display_date.year, self.current_display_date.month, day)
                    
                    bg, fg, font, relief = "white", "black", ("微软雅黑", 10), tk.FLAT
                    
                    if start_d <= current_d <= end_d:
                        bg, fg, font, relief = "#007ACC", "white", ("微软雅黑", 10, "bold"), tk.SOLID
                    elif current_d == today:
                        bg, fg, font, relief = "#E3F2FD", "#1976D2", ("微软雅黑", 10, "bold"), tk.SOLID
                        
                    btn.config(bg=bg, fg=fg, font=font, relief=relief)
                    btn.config(command=lambda w=week_idx, d=day_idx: self.select_day(w, d))
            else:
                for day_idx in range(7):
                    btn = self.day_buttons[week_idx][day_idx]
                    btn.config(text="", state=tk.DISABLED, bg="SystemButtonFace", relief=tk.FLAT)

    def select_day(self, week, day_idx):
        selected_date = datetime(
            self.current_display_date.year,
            self.current_display_date.month,
            int(self.day_buttons[week][day_idx]['text'])
        )
        if self.selecting_start:
            self.start_date = selected_date
            self.end_date = selected_date
            self.selecting_start = False
            self.status_label.config(text="请点击日历选择结束日期", foreground="orange")
        else:
            if selected_date < self.start_date:
                self.end_date = self.start_date
                self.start_date = selected_date
            else:
                self.end_date = selected_date
            self.selecting_start = True
            self.status_label.config(text="时间范围选择完成，可以点击确定", foreground="green")
        self.update_display()

    def prev_month(self):
        current_month_start = self.current_display_date.replace(day=1)
        self.current_display_date = current_month_start - timedelta(days=1)
        self.update_calendar()

    def next_month(self):
        days_in_month = calendar.monthrange(self.current_display_date.year, self.current_display_date.month)[1]
        self.current_display_date = self.current_display_date.replace(day=days_in_month) + timedelta(days=1)
        self.update_calendar()

    def show_year_selector(self, event=None):
        win = tk.Toplevel(self)
        win.title("选择年份")
        win.transient(self)
        win.grab_set()
        x, y = self.year_label.winfo_rootx(), self.year_label.winfo_rooty()
        win.geometry(f"200x300+{x}+{y+30}")
        lb = tk.Listbox(win, font=("微软雅黑", 11), height=15)
        lb.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        cy = self.current_display_date.year
        years = list(range(cy - 10, cy + 11))
        for year in years:
            lb.insert(tk.END, year)
        lb.selection_set(years.index(cy))
        lb.see(years.index(cy))

        def on_select(e=None):
            if lb.curselection():
                self.current_display_date = self.current_display_date.replace(year=years[lb.curselection()[0]])
                self.update_calendar()
            win.destroy()

        lb.bind("<Double-Button-1>", on_select)
        ttk.Button(win, text="确定", command=on_select).pack(pady=(0, 10))

    def update_display(self):
        self.start_date_label.config(text=self.start_date.strftime("%Y-%m-%d"))
        self.end_date_label.config(text=self.end_date.strftime("%Y-%m-%d"))
        self.current_display_date = self.start_date
        self.update_calendar()

    def reset_selection(self):
        self.selecting_start = True
        self.status_label.config(text="请点击日历选择开始日期", foreground="green")

    def _set_range(self, start, end, msg):
        self.start_date, self.end_date = start, end
        self.selecting_start = True
        self.status_label.config(text=msg, foreground="green")
        self.update_display()

    def set_today(self): self._set_range(self.reference_date, self.reference_date, "已选择今天")
    def set_yesterday(self): self._set_range(self.reference_date - timedelta(1), self.reference_date - timedelta(1), "已选择昨天")
    def set_this_week(self): self._set_range(self.reference_date - timedelta(self.reference_date.weekday()), self.reference_date, "已选择本周")
    def set_last_week(self): end = self.reference_date - timedelta(self.reference_date.weekday() + 1); self._set_range(end - timedelta(6), end, "已选择上周")
    def set_this_month(self): self._set_range(self.reference_date.replace(day=1), self.reference_date, "已选择本月")
    def set_last_month(self): end = self.reference_date.replace(day=1) - timedelta(1); self._set_range(end.replace(day=1), end, "已选择上月")
    def set_last_7_days(self): self._set_range(self.reference_date - timedelta(6), self.reference_date, "已选择最近7天")
    def set_last_30_days(self): self._set_range(self.reference_date - timedelta(29), self.reference_date, "已选择最近30天")
    def set_this_quarter(self): qsm = 3 * ((self.reference_date.month - 1) // 3) + 1; self._set_range(self.reference_date.replace(month=qsm, day=1), self.reference_date, "已选择本季度")
    def set_last_quarter(self): end = self.reference_date.replace(month=3 * ((self.reference_date.month - 1) // 3) + 1, day=1) - timedelta(1); self._set_range(end.replace(month=end.month - 2, day=1), end, "已选择上季度")
    def reset_to_reference(self): self._set_range(self.reference_date, self.reference_date, "请点击日历选择开始日期")

    def confirm(self):
        if self.callback:
            self.callback(self.start_date, self.end_date)
        self.destroy()

class DataStatusManager:
    def __init__(self):
        self.status_info = {k: {'last_update': None, 'record_count': 0, 'file_exists': False} for k in
                            ['sales', 'inventory_flow', 'product', 'inventory_check']}
        self.recent_30_days_stats = {
            'total_orders': 0, 'total_quantity': 0, 'total_amount': 0.0, 'date_range': ''
        }

    def update_all_statuses(self, data_frames, product_file_path):
        self.update_sales_status(data_frames.get('sales'))
        self.update_inventory_flow_status(data_frames.get('inventory_flow'))
        self.update_inventory_check_status(data_frames.get('inventory_check'))
        self.update_product_status(product_file_path)

    def update_sales_status(self, sales_df):
        C = Config.STD_COLS
        if sales_df is not None and not sales_df.empty:
            self.status_info['sales'].update({
                'last_update': sales_df[C['SALES_TIME']].max(),
                'record_count': len(sales_df),
                'file_exists': True
            })
            self.calculate_recent_30_days_stats(sales_df)
        else:
            self.status_info['sales']['file_exists'] = False

    def update_inventory_flow_status(self, flow_df):
        if flow_df is not None and not flow_df.empty and '日期' in flow_df.columns:
            self.status_info['inventory_flow'].update({
                'last_update': flow_df['日期'].max(),
                'record_count': len(flow_df),
                'file_exists': True
            })
        else:
            self.status_info['inventory_flow']['file_exists'] = False

    def update_product_status(self, product_file_path):
        if product_file_path and os.path.exists(product_file_path):
            try:
                count = pd.read_excel(product_file_path, usecols=[0]).shape[0]
            except Exception:
                count = 'N/A'
            self.status_info['product'].update({
                'last_update': datetime.fromtimestamp(os.path.getmtime(product_file_path)),
                'record_count': count,
                'file_exists': True
            })
        else:
            self.status_info['product']['file_exists'] = False

    def update_inventory_check_status(self, check_df):
        if check_df is not None and not check_df.empty and '日期' in check_df.columns:
            self.status_info['inventory_check'].update({
                'last_update': check_df['日期'].max(),
                'record_count': len(check_df),
                'file_exists': True
            })
        else:
            self.status_info['inventory_check']['file_exists'] = False

    def calculate_recent_30_days_stats(self, sales_df):
        C = Config.STD_COLS
        try:
            if sales_df is None or sales_df.empty:
                return
            latest_sale_date = sales_df[C['SALES_TIME']].max().date()
            start_date = latest_sale_date - timedelta(days=29)
            recent_sales = sales_df[sales_df[C['SALES_TIME']].dt.date.between(start_date, latest_sale_date)]
            if not recent_sales.empty:
                self.recent_30_days_stats = {
                    'total_orders': recent_sales[C['ORDER_ID']].nunique(),
                    'total_quantity': recent_sales[C['SALES_QTY']].sum(),
                    'total_amount': recent_sales[C['REVENUE']].sum(),
                    'date_range': f"{start_date.strftime('%Y/%m/%d')} ~ {latest_sale_date.strftime('%Y/%m/%d')}"
                }
        except Exception as e:
            print(f"❌ 计算近30天统计失败: {e}")

    def get_status_display_text(self):
        lines = []
        status_map = {
            '销售数据': 'sales', '货流数据': 'inventory_flow',
            '商品资料': 'product', '盘点数据': 'inventory_check'
        }
        for name, key in status_map.items():
            info = self.status_info[key]
            if info['file_exists']:
                update_str = info['last_update'].strftime('%Y-%m-%d %H:%M') if info['last_update'] else 'N/A'
                count_str = f"{info['record_count']:,}" if isinstance(info['record_count'], int) else info['record_count']
                lines.append(f"✅ {name}: 最新至 {update_str} ({count_str}条)")
            else:
                lines.append(f"❌ {name}: 未找到或加载失败")

        stats = self.recent_30_days_stats
        if stats['date_range']:
            lines.append("\n" + "─" * 50)
            lines.append(
                f"📊 近30天 ({stats['date_range']}) 销售概览:\n   总订单: {stats['total_orders']:,} 笔 | 总销量: {stats['total_quantity']:,.0f} 件 | 总金额: {stats['total_amount']:,.2f} 元")
        return '\n'.join(lines)

# ==================== 主入口 ====================
if __name__ == "__main__":
    root = tk.Tk()
    app = SupplierReportGUI(root)
    root.mainloop()
